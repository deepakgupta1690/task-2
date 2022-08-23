#from re import M
import openpyxl
import pandas as pd

path1 = openpyxl.load_workbook(r"C:\Users\Dinesh\Downloads\Input_1.xlsx")
df1 = path1["in"]

new_book = openpyxl.Workbook()
Output = new_book.active

# Adding header

Output["A1"].value ="Name"
Output["B1"].value ="Username"
Output["C1"].value ="Chapter Tag"
Output["D1"].value ="Test_Name"
Output["E1"].value ="answered"
Output["F1"].value ="correct"
Output["G1"].value ="score"
Output["H1"].value ="skipped"
Output["I1"].value ="time-taken (seconds)"
Output["J1"].value ="wrong"

mc = df1.max_column
mr = df1.max_row

# adding testname

for i in range(2,119,9):
    Output.cell(row=i,column=4).value="Concept Test1"
for i in range(3,119,9):
    Output.cell(row=i,column=4).value="Concept Test2"
for i in range(4,119,9):
    Output.cell(row=i,column=4).value="Concept Test3"
for i in range(5,119,9):
    Output.cell(row=i,column=4).value="Concept Test4"
for i in range(6,119,9):
    Output.cell(row=i,column=4).value="Topic Test1"
for i in range(7,119,9):
    Output.cell(row=i,column=4).value="Concept Test5"
for i in range(8,119,9):
    Output.cell(row=i,column=4).value="Topic Test2"
for i in range(9,119,9):
    Output.cell(row=i,column=4).value="Full Chapter Test1"
for i in range(10,119,9):
    Output.cell(row=i,column=4).value="Full Chapter Test1"

# Adding Student Name,Id and Chapter Tag

for x in range(2,15): 
    z = df1.cell(row=x,column=1)
    a = df1.cell(row=x,column=2)
    b = df1.cell(row=x,column=3)
    if x == 2:
        for j in range(2,11):
            Output.cell(row=j,column=1).value=z.value
            Output.cell(row=j,column=2).value=a.value
            Output.cell(row=j,column=3).value=b.value
    if x == 3:
        for j in range(11,20):
            Output.cell(row=j,column=1).value=z.value
            Output.cell(row=j,column=2).value=a.value
            Output.cell(row=j,column=3).value=b.value
    if x == 4:
        for j in range(20,29):
            Output.cell(row=j,column=1).value=z.value
            Output.cell(row=j,column=2).value=a.value
            Output.cell(row=j,column=3).value=b.value
    if x == 5:
        for j in range(29,38):
            Output.cell(row=j,column=1).value=z.value
            Output.cell(row=j,column=2).value=a.value
            Output.cell(row=j,column=3).value=b.value
    if x == 6:
        for j in range(38,47):
            Output.cell(row=j,column=1).value=z.value
            Output.cell(row=j,column=2).value=a.value
            Output.cell(row=j,column=3).value=b.value
    if x == 7:
        for j in range(47,56):
            Output.cell(row=j,column=1).value=z.value
            Output.cell(row=j,column=2).value=a.value
            Output.cell(row=j,column=3).value=b.value
    if x == 8:
        for j in range(56,65):
            Output.cell(row=j,column=1).value=z.value
            Output.cell(row=j,column=2).value=a.value
            Output.cell(row=j,column=3).value=b.value
    if x == 9:
        for j in range(65,74):
            Output.cell(row=j,column=1).value=z.value
            Output.cell(row=j,column=2).value=a.value
            Output.cell(row=j,column=3).value=b.value
    if x == 10:
        for j in range(74,83):
            Output.cell(row=j,column=1).value=z.value
            Output.cell(row=j,column=2).value=a.value
            Output.cell(row=j,column=3).value=b.value
    if x == 11:
        for j in range(83,92):
            Output.cell(row=j,column=1).value=z.value
            Output.cell(row=j,column=2).value=a.value
            Output.cell(row=j,column=3).value=b.value
    if x == 12:
        for j in range(92,101):
            Output.cell(row=j,column=1).value=z.value
            Output.cell(row=j,column=2).value=a.value
            Output.cell(row=j,column=3).value=b.value
    if x == 13:
        for j in range(101,110):
            Output.cell(row=j,column=1).value=z.value
            Output.cell(row=j,column=2).value=a.value
            Output.cell(row=j,column=3).value=b.value
    if x == 14:
        for j in range(110,119):
            Output.cell(row=j,column=1).value=z.value
            Output.cell(row=j,column=2).value=a.value
            Output.cell(row=j,column=3).value=b.value
   
# Adding Answer

j=2
j2=11
j3=20
j4=29
j5=38
j6=47
j7=56
j8=65
j9=74
j10=83
j11=92
j12=101
j13=110       
for i in range (6,mc+1,6):
    c = df1.cell(row=2,column=i)
    d = df1.cell(row=3,column=i)
    e = df1.cell(row=4,column=i)
    f = df1.cell(row=5,column=i)
    g = df1.cell(row=6,column=i)
    h = df1.cell(row=7,column=i)
    k = df1.cell(row=8,column=i)
    l = df1.cell(row=9,column=i)
    m = df1.cell(row=10,column=i)
    n = df1.cell(row=11,column=i)
    o = df1.cell(row=12,column=i)
    p = df1.cell(row=13,column=i)
    q = df1.cell(row=14,column=i)
    Output.cell(row=j,column=5).value=c.value
    Output.cell(row=j2,column=5).value=d.value
    Output.cell(row=j3,column=5).value=e.value
    Output.cell(row=j4,column=5).value=f.value
    Output.cell(row=j5,column=5).value=g.value
    Output.cell(row=j6,column=5).value=h.value
    Output.cell(row=j7,column=5).value=k.value
    Output.cell(row=j8,column=5).value=l.value
    Output.cell(row=j9,column=5).value=m.value
    Output.cell(row=j10,column=5).value=n.value
    Output.cell(row=j11,column=5).value=o.value
    Output.cell(row=j12,column=5).value=p.value
    Output.cell(row=j13,column=5).value=q.value
    j=j+1
    j2=j2+1
    j3=j3+1
    j4=j4+1
    j5=j5+1
    j6=j6+1
    j7=j7+1
    j8=j8+1
    j9=j9+1
    j10=j10+1
    j11=j11+1
    j12=j12+1
    j13=j13+1
      
# Adding correct

j=2
j2=11
j3=20
j4=29
j5=38
j6=47
j7=56
j8=65
j9=74
j10=83
j11=92
j12=101
j13=110       
for i in range (7,mc+1,6):
    c = df1.cell(row=2,column=i)
    d = df1.cell(row=3,column=i)
    e = df1.cell(row=4,column=i)
    f = df1.cell(row=5,column=i)
    g = df1.cell(row=6,column=i)
    h = df1.cell(row=7,column=i)
    k = df1.cell(row=8,column=i)
    l = df1.cell(row=9,column=i)
    m = df1.cell(row=10,column=i)
    n = df1.cell(row=11,column=i)
    o = df1.cell(row=12,column=i)
    p = df1.cell(row=13,column=i)
    q = df1.cell(row=14,column=i)
    Output.cell(row=j,column=6).value=c.value
    Output.cell(row=j2,column=6).value=d.value
    Output.cell(row=j3,column=6).value=e.value
    Output.cell(row=j4,column=6).value=f.value
    Output.cell(row=j5,column=6).value=g.value
    Output.cell(row=j6,column=6).value=h.value
    Output.cell(row=j7,column=6).value=k.value
    Output.cell(row=j8,column=6).value=l.value
    Output.cell(row=j9,column=6).value=m.value
    Output.cell(row=j10,column=6).value=n.value
    Output.cell(row=j11,column=6).value=o.value
    Output.cell(row=j12,column=6).value=p.value
    Output.cell(row=j13,column=6).value=q.value
    j=j+1
    j2=j2+1
    j3=j3+1
    j4=j4+1
    j5=j5+1
    j6=j6+1
    j7=j7+1
    j8=j8+1
    j9=j9+1
    j10=j10+1
    j11=j11+1
    j12=j12+1
    j13=j13+1

# Adding wrong

j=2
j2=11
j3=20
j4=29
j5=38
j6=47
j7=56
j8=65
j9=74
j10=83
j11=92
j12=101
j13=110               
for i in range (8,mc+1,6):
    c = df1.cell(row=2,column=i)
    d = df1.cell(row=3,column=i)
    e = df1.cell(row=4,column=i)
    f = df1.cell(row=5,column=i)
    g = df1.cell(row=6,column=i)
    h = df1.cell(row=7,column=i)
    k = df1.cell(row=8,column=i)
    l = df1.cell(row=9,column=i)
    m = df1.cell(row=10,column=i)
    n = df1.cell(row=11,column=i)
    o = df1.cell(row=12,column=i)
    p = df1.cell(row=13,column=i)
    q = df1.cell(row=14,column=i)
    Output.cell(row=j,column=10).value=c.value
    Output.cell(row=j2,column=10).value=d.value
    Output.cell(row=j3,column=10).value=e.value
    Output.cell(row=j4,column=10).value=f.value
    Output.cell(row=j5,column=10).value=g.value
    Output.cell(row=j6,column=10).value=h.value
    Output.cell(row=j7,column=10).value=k.value
    Output.cell(row=j8,column=10).value=l.value
    Output.cell(row=j9,column=10).value=m.value
    Output.cell(row=j10,column=10).value=n.value
    Output.cell(row=j11,column=10).value=o.value
    Output.cell(row=j12,column=10).value=p.value
    Output.cell(row=j13,column=10).value=q.value
    j=j+1
    j2=j2+1
    j3=j3+1
    j4=j4+1
    j5=j5+1
    j6=j6+1
    j7=j7+1
    j8=j8+1
    j9=j9+1
    j10=j10+1
    j11=j11+1
    j12=j12+1
    j13=j13+1

# Adding skipped

j=2
j2=11
j3=20
j4=29
j5=38
j6=47
j7=56
j8=65
j9=74
j10=83
j11=92
j12=101
j13=110          
for i in range (9,mc+1,6):
    c = df1.cell(row=2,column=i)
    d = df1.cell(row=3,column=i)
    e = df1.cell(row=4,column=i)
    f = df1.cell(row=5,column=i)
    g = df1.cell(row=6,column=i)
    h = df1.cell(row=7,column=i)
    k = df1.cell(row=8,column=i)
    l = df1.cell(row=9,column=i)
    m = df1.cell(row=10,column=i)
    n = df1.cell(row=11,column=i)
    o = df1.cell(row=12,column=i)
    p = df1.cell(row=13,column=i)
    q = df1.cell(row=14,column=i)
    Output.cell(row=j,column=8).value=c.value
    Output.cell(row=j2,column=8).value=d.value
    Output.cell(row=j3,column=8).value=e.value
    Output.cell(row=j4,column=8).value=f.value
    Output.cell(row=j5,column=8).value=g.value
    Output.cell(row=j6,column=8).value=h.value
    Output.cell(row=j7,column=8).value=k.value
    Output.cell(row=j8,column=8).value=l.value
    Output.cell(row=j9,column=8).value=m.value
    Output.cell(row=j10,column=8).value=n.value
    Output.cell(row=j11,column=8).value=o.value
    Output.cell(row=j12,column=8).value=p.value
    Output.cell(row=j13,column=8).value=q.value
    j=j+1
    j2=j2+1
    j3=j3+1
    j4=j4+1
    j5=j5+1
    j6=j6+1
    j7=j7+1
    j8=j8+1
    j9=j9+1
    j10=j10+1
    j11=j11+1
    j12=j12+1
    j13=j13+1

# Adding score

j=2
j2=11
j3=20
j4=29
j5=38
j6=47
j7=56
j8=65
j9=74
j10=83
j11=92
j12=101
j13=110        
for i in range (4,mc+1,6):
    c = df1.cell(row=2,column=i)
    d = df1.cell(row=3,column=i)
    e = df1.cell(row=4,column=i)
    f = df1.cell(row=5,column=i)
    g = df1.cell(row=6,column=i)
    h = df1.cell(row=7,column=i)
    k = df1.cell(row=8,column=i)
    l = df1.cell(row=9,column=i)
    m = df1.cell(row=10,column=i)
    n = df1.cell(row=11,column=i)
    o = df1.cell(row=12,column=i)
    p = df1.cell(row=13,column=i)
    q = df1.cell(row=14,column=i)
    Output.cell(row=j,column=7).value=c.value
    Output.cell(row=j2,column=7).value=d.value
    Output.cell(row=j3,column=7).value=e.value
    Output.cell(row=j4,column=7).value=f.value
    Output.cell(row=j5,column=7).value=g.value
    Output.cell(row=j6,column=7).value=h.value
    Output.cell(row=j7,column=7).value=k.value
    Output.cell(row=j8,column=7).value=l.value
    Output.cell(row=j9,column=7).value=m.value
    Output.cell(row=j10,column=7).value=n.value
    Output.cell(row=j11,column=7).value=o.value
    Output.cell(row=j12,column=7).value=p.value
    Output.cell(row=j13,column=7).value=q.value
    j=j+1
    j2=j2+1
    j3=j3+1
    j4=j4+1
    j5=j5+1
    j6=j6+1
    j7=j7+1
    j8=j8+1
    j9=j9+1
    j10=j10+1
    j11=j11+1
    j12=j12+1
    j13=j13+1

# Adding timetaken

j=2 
j2=11
j3=20
j4=29
j5=38
j6=47
j7=56
j8=65
j9=74
j10=83
j11=92
j12=101
j13=110            
for i in range (5,mc+1,6):
    c = df1.cell(row=2,column=i)
    d = df1.cell(row=3,column=i)
    e = df1.cell(row=4,column=i)
    f = df1.cell(row=5,column=i)
    g = df1.cell(row=6,column=i)
    h = df1.cell(row=7,column=i)
    k = df1.cell(row=8,column=i)
    l = df1.cell(row=9,column=i)
    m = df1.cell(row=10,column=i)
    n = df1.cell(row=11,column=i)
    o = df1.cell(row=12,column=i)
    p = df1.cell(row=13,column=i)
    q = df1.cell(row=14,column=i)
    Output.cell(row=j,column=9).value=c.value
    Output.cell(row=j2,column=9).value=d.value
    Output.cell(row=j3,column=9).value=e.value
    Output.cell(row=j4,column=9).value=f.value
    Output.cell(row=j5,column=9).value=g.value
    Output.cell(row=j6,column=9).value=h.value
    Output.cell(row=j7,column=9).value=k.value
    Output.cell(row=j8,column=9).value=l.value
    Output.cell(row=j9,column=9).value=m.value
    Output.cell(row=j10,column=9).value=n.value
    Output.cell(row=j11,column=9).value=o.value
    Output.cell(row=j12,column=9).value=p.value
    Output.cell(row=j13,column=9).value=q.value
    j=j+1
    j2=j2+1
    j3=j3+1
    j4=j4+1
    j5=j5+1
    j6=j6+1
    j7=j7+1
    j8=j8+1
    j9=j9+1
    j10=j10+1
    j11=j11+1
    j12=j12+1
    j13=j13+1

new_book.save(r"C:\Users\Dinesh\Downloads\task2.xlsx")